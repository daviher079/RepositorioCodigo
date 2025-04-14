import pandas as pd
import openpyxl
from openpyxl import load_workbook
from collections import Counter

ruta_fichero = 'PruebasPandas/ControlEntrenamientos24-25.xlsx'
workbook = load_workbook(ruta_fichero)
hoja = workbook['Hoja1']

n_filas = hoja.max_row
n_columnas = 9

# df = pd.read_excel(ruta_fichero, sheet_name='Hoja1')

colores_por_fila = []

for fila in range(2, n_filas + 1):
    colores_fila = []
    for columna in range(1, n_columnas + 1):
        celda = hoja.cell(row=fila, column=columna)
        if celda.fill.fgColor.rgb != '00000000':
            color = celda.fill.fgColor.rgb
        else:
            color = celda.value
        colores_fila.append(color)
    colores_por_fila.append(colores_fila)    

columnas = [f'Columna_{i+1}' for i in range(n_columnas)]

df_nombres = pd.DataFrame(colores_por_fila, columns=columnas)

columnas_a_mantener = []
for col in columnas:
    if df_nombres[col][1:].notna().any():
        columnas_a_mantener.append(col)

df_nombres = df_nombres[columnas_a_mantener]

leyenda = {
    'FFFF0000': 'Nº de días que no avisó',
    'FF6AA84F': 'Nº de días que entrenó',
    'FF7030A0': 'Nº de días que avisó y no entrenó',
}

salida_real = df_nombres.replace(leyenda)

datos_jugadores = []
for index, row in salida_real.iterrows():
    # Contar valores excluyendo None con dropna()
    conteo = Counter(row.dropna())
    datos_jugadores.append(conteo)

nombres = [list(counter.keys())[list(counter.values()).index(1)] for counter in datos_jugadores]

datos_procesados = {
    'Nombre': nombres,
    'Días entrenados': [counter['Nº de días que entrenó'] for counter in datos_jugadores],
    'Días avisó y no entrenó': [counter.get('Nº de días que avisó y no entrenó', 0) for counter in datos_jugadores],
    'Días no avisó': [counter.get('Nº de días que no avisó', 0) for counter in datos_jugadores]
}

# Crear el DataFrame
df = pd.DataFrame(datos_procesados)

# Ordenar por 'Días entrenados' de mayor a menor
df_ordenado = df.sort_values(by='Días entrenados', ascending=False)

# Mostrar el resultado
print(df_ordenado)